"""
evaluation/excel_report.py
───────────────────────────
Pembuat laporan Excel (.xlsx) hasil evaluasi — revisi pasca-sidang item 6.

Satu file eval_YYYYMMDD_HHMMSS.xlsx berisi delapan sheet:

  1. Ringkasan          — konfigurasi lengkap (model, K, θ, θc, bobot) +
                          agregat A vs B + status hipotesis.
  2. Metrik Per Kasus   — satu baris per kasus uji: seluruh metrik retrieval
                          & generasi kedua kondisi + delta.
  3. Perhitungan Detail — SETIAP perhitungan metrik direkam eksplisit:
                          rumus, substitusi angka, hasil (per kasus, per
                          metrik, per kondisi).
  4. Respons A vs B     — query utuh + respons LLM UTUH kedua kondisi
                          berdampingan agar mudah dibandingkan.
  5. Chunk & Sumber     — per kasus per peringkat: NAMA FILE, TOPIK, skor,
                          lolos θ (0,25) dan θc (0,35), cuplikan teks chunk.
  6. Uncertainty-Contradiction — setiap frasa yang terdeteksi: di kondisi
                          mana, frasa apa, KALIMAT KE BERAPA, POSISI KARAKTER,
                          beserta kutipan kalimatnya.
  7. Entailment Detail  — verdict YA/TIDAK per klaim (LLM-as-Judge) untuk
                          kedua kondisi — dasar perhitungan Faithfulness.
  8. Leksikon           — daftar frasa Uncertainty & Contradiction versi
                          revisi (bahan pembaruan Lampiran 5).

Dipanggil dari evaluation/runner.py → save_results().
"""

import logging
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from evaluation.lexicon import lexicon_table

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SUB_FILL    = PatternFill("solid", fgColor="DDEBF7")
_WRAP        = Alignment(wrap_text=True, vertical="top")
_TOP         = Alignment(vertical="top")


def _sheet_with_header(wb: Workbook, title: str, headers: List[str],
                       widths: Optional[List[int]] = None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for col_i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_i)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        if widths and col_i <= len(widths):
            ws.column_dimensions[get_column_letter(col_i)].width = widths[col_i - 1]
    ws.freeze_panes = "A2"
    return ws


def _wrap_row(ws, row_idx: int, cols: List[int]) -> None:
    for c in cols:
        ws.cell(row=row_idx, column=c).alignment = _WRAP


def _g(d: Optional[Dict], *keys, default=None):
    """Ambil nilai bersarang dengan aman: _g(r, 'kondisi_a', 'faithfulness')."""
    cur: Any = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
    return cur if cur is not None else default


# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — RINGKASAN
# ══════════════════════════════════════════════════════════════════════════

def _sheet_ringkasan(wb: Workbook, aggregates: Dict, config: Dict) -> None:
    ws = wb.active
    ws.title = "Ringkasan"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30

    def head(text: str):
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL

    head("KONFIGURASI EVALUASI (revisi pasca-sidang)")
    for k, v in config.items():
        ws.append([k, str(v)])
    ws.append([])

    head("METRIK RETRIEVAL — KONDISI A (Pers. 2, 4, 6, 8, 10)")
    ret = aggregates.get("retrieval", {})
    for label, key in [
        ("Precision@K (θ=0,25)",  "avg_precision_at_k"),
        ("Recall@K (θ=0,25)",     "avg_recall_at_k"),
        ("Mean Similarity",        "avg_mean_similarity"),
        ("Coverage (θc=0,35)",    "avg_coverage"),
        ("Source Diversity",       "avg_source_diversity"),
    ]:
        ws.append([label, ret.get(key)])
    ws.append([])

    head("PERBANDINGAN GENERASI — KONDISI A vs B")
    ws.append(["Metrik", "Kondisi A (RAG)", "Kondisi B (tanpa RAG)",
               "Delta (A-B)", "Hipotesis"])
    for c in range(1, 6):
        ws.cell(row=ws.max_row, column=c).fill = _SUB_FILL
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)

    ka, kb = aggregates.get("kondisi_a", {}), aggregates.get("kondisi_b", {})
    kom = aggregates.get("komparasi", {})
    fa, fb = ka.get("avg_faithfulness"), kb.get("avg_faithfulness")
    ha, hb = ka.get("avg_hallucination_risk"), kb.get("avg_hallucination_risk")
    aa = _g(ka, "answer_accuracy", "accuracy")
    ab = _g(kb, "answer_accuracy", "accuracy")

    def verdict(cond: Optional[bool], hyp: str) -> str:
        if cond is None:
            return hyp
        return f"{hyp} ({'TERPENUHI' if cond else 'TIDAK TERPENUHI'})"

    ws.append(["Faithfulness", fa, fb, kom.get("delta_faithfulness"),
               verdict(None if fa is None or fb is None else fa > fb, "A > B")])
    ws.append(["Hallucination Risk", ha, hb, kom.get("delta_hallucination_risk"),
               verdict(None if ha is None or hb is None else ha < hb, "A < B")])
    ws.append(["Answer Accuracy", aa, ab, kom.get("delta_answer_accuracy"),
               verdict(None if aa is None or ab is None else aa > ab, "A > B")])
    ws.append([])

    head("DISTRIBUSI RISIKO HALUSINASI")
    ws.append(["Kategori", "A (n)", "A (%)", "B (n)", "B (%)"])
    hda, hdb = ka.get("hallucination_dist", {}), kb.get("hallucination_dist", {})
    for cat in ["rendah", "sedang", "tinggi"]:
        da, db = hda.get(cat, {}), hdb.get(cat, {})
        ws.append([cat.upper(), da.get("n"), da.get("pct"),
                   db.get("n"), db.get("pct")])
    ws.append([])

    head("EFFECT SIZE & PERBAIKAN RELATIF")
    ws.append(["Cohen's d (Faithfulness)", kom.get("cohen_d_faithfulness")])
    ws.append(["Faithfulness improvement (%)", kom.get("relative_faith_improvement_pct")])
    ws.append(["Hallucination reduction (%)", kom.get("relative_hall_reduction_pct")])
    ws.append(["Jumlah kasus uji dievaluasi", aggregates.get("n_tested")])


# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — METRIK PER KASUS
# ══════════════════════════════════════════════════════════════════════════

def _sheet_metrik(wb: Workbook, results: List[Dict]) -> None:
    ws = _sheet_with_header(wb, "Metrik Per Kasus", [
        "test_id", "kategori", "kognitif", "query (cuplikan)",
        "P@K", "R@K", "MeanSim", "Coverage", "SourceDiv",
        "sumber unik (Diversity)",
        "A_Faithfulness", "A_Entailment", "A_KWoverlap",
        "A_HallRisk", "A_Label", "A_Contradiction", "A_Uncertainty", "A_Benar",
        "B_Faithfulness", "B_Entailment", "B_KWoverlap",
        "B_HallRisk", "B_Label", "B_Contradiction", "B_Uncertainty", "B_Benar",
        "ΔFaithfulness", "ΔHallRisk",
    ], widths=[7, 12, 8, 46, 8, 8, 9, 9, 9, 40,
               13, 12, 12, 10, 9, 13, 12, 8,
               13, 12, 12, 10, 9, 13, 12, 8, 13, 10])
    for r in results:
        ret = r.get("retrieval", {})
        fa = _g(r, "kondisi_a", "faithfulness") or {}
        fb = _g(r, "kondisi_b", "faithfulness") or {}
        ha = _g(r, "kondisi_a", "hallucination") or {}
        hb = _g(r, "kondisi_b", "hallucination") or {}
        div_detail = ret.get("source_diversity_detail") or {}
        d_f = (round(fa.get("faithfulness_score") - fb.get("faithfulness_score"), 4)
               if fa.get("faithfulness_score") is not None
               and fb.get("faithfulness_score") is not None else None)
        d_h = (round(ha.get("hallucination_risk") - hb.get("hallucination_risk"), 4)
               if ha.get("hallucination_risk") is not None
               and hb.get("hallucination_risk") is not None else None)
        ws.append([
            r.get("test_id"), r.get("query_type"), r.get("cognitive"),
            (r.get("query") or "")[:120],
            ret.get("precision_at_k"), ret.get("recall_at_k"),
            ret.get("mean_similarity"), ret.get("coverage"),
            ret.get("source_diversity"),
            ", ".join(div_detail.get("unique_sources", [])),
            fa.get("faithfulness_score"), fa.get("entailment_score"),
            fa.get("keyword_overlap"),
            ha.get("hallucination_risk"), ha.get("risk_label"),
            ha.get("contradiction_score"),
            ha.get("uncertainty_value"),
            _g(r, "kondisi_a", "answer_correct"),
            fb.get("faithfulness_score"), fb.get("entailment_score"),
            fb.get("keyword_overlap"),
            hb.get("hallucination_risk"), hb.get("risk_label"),
            hb.get("contradiction_score"),
            hb.get("uncertainty_value"),
            _g(r, "kondisi_b", "answer_correct"),
            d_f, d_h,
        ])
        _wrap_row(ws, ws.max_row, [4, 10])


# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — PERHITUNGAN DETAIL (setiap langkah metrik)
# ══════════════════════════════════════════════════════════════════════════

def _sheet_perhitungan(wb: Workbook, results: List[Dict]) -> None:
    ws = _sheet_with_header(wb, "Perhitungan Detail", [
        "test_id", "kondisi", "metrik", "rumus", "substitusi / langkah",
        "hasil", "keterangan",
    ], widths=[7, 8, 22, 46, 60, 10, 46])

    def add(tid, kondisi, metrik, rumus, subst, hasil, ket=""):
        ws.append([tid, kondisi, metrik, rumus, subst, hasil, ket])
        _wrap_row(ws, ws.max_row, [4, 5, 7])

    for r in results:
        tid = r.get("test_id")
        ret = r.get("retrieval", {})

        for key, name in [
            ("precision_detail",  "Precision@K (Pers. 2)"),
            ("recall_detail",     "Recall@K (Pers. 4)"),
            ("mean_sim_detail",   "Mean Similarity (Pers. 6)"),
            ("coverage_detail",   "Coverage (Pers. 8)"),
            ("source_diversity_detail", "Source Diversity (Pers. 10)"),
        ]:
            d = ret.get(key)
            if d:
                ket = ""
                if key == "precision_detail":
                    ket = f"basis: {ret.get('precision_basis', 'skor chunk nyata')}"
                if key == "source_diversity_detail":
                    ket = "file unik: " + ", ".join(d.get("unique_sources", []))
                add(tid, "A", name, d.get("formula"),
                    f"skor = {d.get('scores_used', d.get('unique_sources'))} → "
                    f"{d.get('substitution')}",
                    d.get("value"), ket)

        for cond in ("kondisi_a", "kondisi_b"):
            label = "A" if cond == "kondisi_a" else "B"
            f = _g(r, cond, "faithfulness") or {}
            if f.get("faithfulness_score") is not None:
                ent = f.get("entailment_score")
                kwd = f.get("kw_overlap_detail") or {}
                add(tid, label, "Faithfulness (Pers. 15)",
                    "F = 0,70×Entailment + 0,30×KWoverlap",
                    f.get("calc_str", ""),
                    f.get("faithfulness_score"),
                    f"metode: {f.get('method', '')}")
                if ent is not None:
                    add(tid, label, "— Entailment (Pers. 12-13)",
                        "Entailment = Σverdict(sᵢ)/m",
                        f"{f.get('claims_supported')} klaim YA dari "
                        f"{f.get('claims_evaluated')} klaim dievaluasi",
                        ent, "verdict per klaim: sheet 'Entailment Detail'")
                if kwd:
                    add(tid, label, "— Keyword Overlap (Pers. 14)",
                        "KWoverlap = (|Stems(ans)∩Stems(ctx)| / |Stems(ctx)|)^0,65",
                        kwd.get("calc_str", ""), f.get("keyword_overlap"))
            h = _g(r, cond, "hallucination") or {}
            if h.get("hallucination_risk") is not None:
                add(tid, label, "Hallucination Risk (Pers. 18)",
                    "Risk = 0,65×(1-F) + 0,20×Contradiction + 0,15×Uncertainty",
                    h.get("calc_str", ""),
                    h.get("hallucination_risk"),
                    f"label: {h.get('risk_label')}; "
                    f"kontradiksi {h.get('contradiction_n_flagged', 0)}/"
                    f"{h.get('contradiction_n_sentences', 0)} kalimat; "
                    f"uncertainty flag: {h.get('uncertainty_flag')}")
            aq = _g(r, cond, "answer_quality") or {}
            if _g(r, cond, "answer_correct") is not None:
                add(tid, label, "Answer Accuracy (Pers. 20)",
                    "Correctᵢ ∈ {0,1} via LLM-as-Judge (HASIL: BENAR/SALAH)",
                    (aq.get("feedback") or "")[:400],
                    "BENAR" if _g(r, cond, "answer_correct") else "SALAH",
                    f"evaluator: {aq.get('evaluated_by', '')}")


# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — RESPONS LLM A vs B (utuh)
# ══════════════════════════════════════════════════════════════════════════

def _sheet_respons(wb: Workbook, results: List[Dict]) -> None:
    ws = _sheet_with_header(wb, "Respons A vs B", [
        "test_id", "kategori", "kognitif", "query utuh",
        "Kondisi A — respons UTUH (RAG)", "A_Faith", "A_HallRisk", "A_Benar",
        "Kondisi B — respons UTUH (tanpa RAG)", "B_Faith", "B_HallRisk", "B_Benar",
        "ΔFaith",
    ], widths=[7, 12, 8, 45, 70, 9, 10, 8, 70, 9, 10, 8, 9])
    for r in results:
        fa = _g(r, "kondisi_a", "faithfulness") or {}
        fb = _g(r, "kondisi_b", "faithfulness") or {}
        ha = _g(r, "kondisi_a", "hallucination") or {}
        hb = _g(r, "kondisi_b", "hallucination") or {}
        d_f = (round(fa.get("faithfulness_score") - fb.get("faithfulness_score"), 4)
               if fa.get("faithfulness_score") is not None
               and fb.get("faithfulness_score") is not None else None)
        ws.append([
            r.get("test_id"), r.get("query_type"), r.get("cognitive"),
            r.get("query"),
            _g(r, "kondisi_a", "reply_full") or "",
            fa.get("faithfulness_score"), ha.get("hallucination_risk"),
            _g(r, "kondisi_a", "answer_correct"),
            _g(r, "kondisi_b", "reply_full") or "",
            fb.get("faithfulness_score"), hb.get("hallucination_risk"),
            _g(r, "kondisi_b", "answer_correct"),
            d_f,
        ])
        _wrap_row(ws, ws.max_row, [4, 5, 9])


# ══════════════════════════════════════════════════════════════════════════
# SHEET 5 — CHUNK & SUMBER (nama file + topik per peringkat)
# ══════════════════════════════════════════════════════════════════════════

def _sheet_chunks(wb: Workbook, results: List[Dict]) -> None:
    ws = _sheet_with_header(wb, "Chunk & Sumber", [
        "test_id", "peringkat", "file sumber", "topik", "skor cosine",
        "≥ θ (0,25)", "≥ θc (0,35)", "cuplikan teks chunk",
    ], widths=[7, 9, 40, 44, 12, 10, 10, 80])
    for r in results:
        for row in (r.get("retrieval", {}).get("chunks_detail") or []):
            ws.append([
                r.get("test_id"), row.get("rank"), row.get("source"),
                row.get("topic"), row.get("score"),
                "✓" if row.get("ge_theta") else "✗",
                "✓" if row.get("ge_theta_c") else "✗",
                row.get("preview"),
            ])
            _wrap_row(ws, ws.max_row, [4, 8])


# ══════════════════════════════════════════════════════════════════════════
# SHEET 6 — UNCERTAINTY & CONTRADICTION (lokasi frasa)
# ══════════════════════════════════════════════════════════════════════════

def _sheet_lexmatches(wb: Workbook, results: List[Dict]) -> None:
    ws = _sheet_with_header(wb, "Uncertainty-Contradiction", [
        "test_id", "kondisi", "jenis", "frasa terdeteksi",
        "kalimat ke-", "posisi karakter", "kutipan kalimat",
    ], widths=[7, 8, 14, 26, 11, 15, 90])
    found = False
    for r in results:
        for cond in ("kondisi_a", "kondisi_b"):
            label = "A" if cond == "kondisi_a" else "B"
            h = _g(r, cond, "hallucination") or {}
            for jenis, key in [("Uncertainty", "uncertainty_matches"),
                               ("Contradiction", "contradiction_matches")]:
                for m in h.get(key) or []:
                    found = True
                    ws.append([
                        r.get("test_id"), label, jenis, m.get("phrase"),
                        (m.get("sentence_index", -1) + 1) or "-",
                        f"{m.get('char_start')}–{m.get('char_end')}",
                        m.get("sentence"),
                    ])
                    _wrap_row(ws, ws.max_row, [7])
    if not found:
        ws.append(["-", "-", "-", "(tidak ada frasa terdeteksi pada seluruh kasus)",
                   "-", "-", "-"])


# ══════════════════════════════════════════════════════════════════════════
# SHEET 7 — ENTAILMENT DETAIL (verdict per klaim)
# ══════════════════════════════════════════════════════════════════════════

def _sheet_entailment(wb: Workbook, results: List[Dict]) -> None:
    ws = _sheet_with_header(wb, "Entailment Detail", [
        "test_id", "kondisi", "klaim ke-", "klaim (kalimat respons LLM)",
        "verdict LLM-as-Judge",
    ], widths=[7, 8, 9, 100, 20])
    for r in results:
        for cond in ("kondisi_a", "kondisi_b"):
            label = "A" if cond == "kondisi_a" else "B"
            f = _g(r, cond, "faithfulness") or {}
            for i, d in enumerate(f.get("entailment_detail") or [], 1):
                sup = d.get("supported")
                verdict = ("YA (didukung konteks)" if sup is True
                           else "TIDAK (tidak didukung)" if sup is False
                           else "GAGAL PARSE")
                ws.append([r.get("test_id"), label, i, d.get("claim"), verdict])
                _wrap_row(ws, ws.max_row, [4])


# ══════════════════════════════════════════════════════════════════════════
# SHEET 8 — LEKSIKON (bahan pembaruan Lampiran 5)
# ══════════════════════════════════════════════════════════════════════════

def _sheet_lexicon(wb: Workbook) -> None:
    ws = _sheet_with_header(wb, "Leksikon", [
        "jenis", "frasa", "keterangan",
    ], widths=[16, 30, 80])
    ket = {
        "Uncertainty":   ("Memicu komponen Uncertainty = 0,20 (biner) pada "
                          "Persamaan 18. Leksikon revisi pasca-sidang — lihat "
                          "evaluation/lexicon.py"),
        "Contradiction": ("Dihitung sebagai proporsi kalimat berfrasa "
                          "kontradiksi kuat pada Persamaan 18. Leksikon revisi "
                          "pasca-sidang — lihat evaluation/lexicon.py"),
    }
    for row in lexicon_table():
        ws.append([row["jenis"], row["frasa"], ket[row["jenis"]]])
        _wrap_row(ws, ws.max_row, [3])


# ══════════════════════════════════════════════════════════════════════════
# API PUBLIK
# ══════════════════════════════════════════════════════════════════════════

def build_excel(results: List[Dict], aggregates: Dict, config: Dict,
                out_path: str) -> str:
    """
    Susun seluruh sheet dan simpan workbook ke *out_path*.

    Parameters
    ----------
    results    : list hasil per kasus dari evaluation/runner.py
    aggregates : dict agregat dari compute_aggregates()
    config     : dict konfigurasi evaluasi (model, K, ambang, bobot, dst.)
    out_path   : path file .xlsx tujuan

    Returns
    -------
    str — path file yang tersimpan.
    """
    wb = Workbook()
    _sheet_ringkasan(wb, aggregates, config)
    _sheet_metrik(wb, results)
    _sheet_perhitungan(wb, results)
    _sheet_respons(wb, results)
    _sheet_chunks(wb, results)
    _sheet_lexmatches(wb, results)
    _sheet_entailment(wb, results)
    _sheet_lexicon(wb)
    wb.save(out_path)
    logger.info("Laporan Excel tersimpan: %s", out_path)
    return out_path
